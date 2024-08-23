import os
import pandas as pd
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackContext, ConversationHandler
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Токен бота
TOKEN = '7061979927:AAHZg2lFtJ0v6rGySef6pmRD9LJKitYu3ic'

# Путь к папке "Mayer" на диске D
MAYER_FOLDER_PATH = 'D:\\Bot\\Mayer'
def adjust_column_widths(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active

    # Настройка ширины всех столбцов
    for column in worksheet.columns:
        col_letter = column[0].column_letter
        worksheet.column_dimensions[col_letter].width = 31.0

    # Сохранение изменений в Excel-файле
    workbook.save(file_path)

def adjust_row_heights_and_alignments(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active

    # Настройка высоты всех строк и выравнивания текста
    for row in worksheet.iter_rows():
        worksheet.row_dimensions[row[0].row].height = 30.0  # Пример высоты строки
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    workbook.save(file_path)

# Путь к Excel-файлу
EXCEL_FILE_PATH = os.path.join(MAYER_FOLDER_PATH, 'employees.xlsx')

# Убедитесь, что папка Mayer существует
if not os.path.exists(MAYER_FOLDER_PATH):
    os.makedirs(MAYER_FOLDER_PATH)

# Проверка наличия столбцов в Excel
if not os.path.exists(EXCEL_FILE_PATH):
    df = pd.DataFrame(columns=[
        'Имя Работника', 'Прописка в Украине', 'Прописка в Словакии', 'IBAN', 'Номер телефона', 'Страховка', 
        'Информация о страховке', 'Вторая работа', 'Информация о второй работе', 'Образование'
    ])
    df.to_excel(EXCEL_FILE_PATH, index=False)
    adjust_column_widths(EXCEL_FILE_PATH)
    adjust_row_heights_and_alignments(EXCEL_FILE_PATH)
else:
    df = pd.read_excel(EXCEL_FILE_PATH)
    if 'Прописка в Украине' not in df.columns:
        df['Прописка в Украине'] = ''
    if 'Прописка в Словакии' not in df.columns:
        df['Прописка в Словакии'] = ''
    if 'IBAN' not in df.columns:
        df['IBAN'] = ''
    if 'Номер телефона' not in df.columns:
        df['Номер телефона'] = ''
    if 'Страховка' not in df.columns:
        df['Страховка'] = ''
    if 'Информация о страховке' not in df.columns:
        df['Информация о страховке'] = ''
    if 'Вторая работа' not in df.columns:
        df['Вторая работа'] = ''
    if 'Информация о второй работе' not in df.columns:
        df['Информация о второй работе'] = ''  # Добавляем новый столбец
    if 'Образование' not in df.columns:
        df['Образование'] = ''  # Добавляем новый столбец
    df.to_excel(EXCEL_FILE_PATH, index=False)
    adjust_column_widths(EXCEL_FILE_PATH)
    adjust_row_heights_and_alignments(EXCEL_FILE_PATH)



# Состояния для ConversationHandler
(
    WAITING_FOR_NAME,
    WAITING_FOR_PASSPORT,
    WAITING_FOR_REFUGEE,
    WAITING_FOR_SLOVAKIA_REGISTRATION,
    WAITING_FOR_UKRAINE_REGISTRATION,
    WAITING_FOR_BANK_DETAILS,
    WAITING_FOR_SLOVAK_PHONE,
    WAITING_FOR_INSURANCE,
    WAITING_FOR_INSURANCE_INFO,
    WAITING_FOR_SECOND_JOB_CONFIRMATION,
    WAITING_FOR_SECOND_JOB,
    WAITING_FOR_EDUCATION_CONFIRMATION,
    WAITING_FOR_EDUCATION,
    WAITING_FOR_STUDENT_CARD,
    WAITING_FOR_UNIVERSITY_CONFIRMATION
) = range(15)

# Начало общения
async def start(update: Update, context: CallbackContext) -> int:
    await update.message.reply_text(
        'Добрый день, напишите свое ФИО латыницей!'
    )
    return WAITING_FOR_NAME

# Обработка введенного ФИО
async def handle_name(update: Update, context: CallbackContext) -> int:
    full_name = update.message.text.strip()
    context.user_data['full_name'] = full_name

    # Сохранение ФИО в Excel-файл
    df = pd.read_excel(EXCEL_FILE_PATH)
    new_entry = pd.DataFrame([{
        'Имя Работника': full_name,
        'Прописка в Украине': '',
        'Прописка в Словакии': ''
    }])
    df = pd.concat([df, new_entry], ignore_index=True)
    df.to_excel(EXCEL_FILE_PATH, index=False)
    adjust_column_widths(EXCEL_FILE_PATH)
    adjust_row_heights_and_alignments(EXCEL_FILE_PATH)
    
    # Настройка ширины столбцов после сохранения данных
    adjust_column_widths(EXCEL_FILE_PATH)
    adjust_row_heights_and_alignments(EXCEL_FILE_PATH)

    # Создание папки с именем пользователя в папке Mayer
    user_folder_path = os.path.join(MAYER_FOLDER_PATH, full_name)

    try:
        os.makedirs(user_folder_path, exist_ok=True)
        print(f"Папка создана: {user_folder_path}")
    except Exception as e:
        print(f"Ошибка при создании папки: {e}")

    await update.message.reply_text(
        'Для оформления нам потребуется скан загран. паспорта, отправьте его, пожалуйста.'
    )
    return WAITING_FOR_PASSPORT


# Обработка скана заграничного паспорта
async def handle_passport(update: Update, context: CallbackContext) -> int:
    if update.message.photo or update.message.document:
        user_name = context.user_data.get('full_name', 'unknown_user')
        user_folder_path = os.path.join(MAYER_FOLDER_PATH, user_name)
        
        if not os.path.exists(user_folder_path):
            os.makedirs(user_folder_path, exist_ok=True)

        file_id = update.message.photo[-1].file_id if update.message.photo else update.message.document.file_id
        file = await context.bot.get_file(file_id)
        file_path = os.path.join(user_folder_path, file.file_path.split('/')[-1])
        
        await file.download_to_drive(file_path)
        
        await update.message.reply_text(
            'Прошу Вас отправить скан Беженства/ВНЖ'
        )
        return WAITING_FOR_REFUGEE
    else:
        await update.message.reply_text(
            'Пожалуйста, отправьте скан заграничного паспорта в формате фотографии или документа.'
        )
        return WAITING_FOR_PASSPORT

# Обработка скана беженства/ВНЖ
async def handle_refugee(update: Update, context: CallbackContext) -> int:
    if update.message.photo or update.message.document:
        user_name = context.user_data.get('full_name', 'unknown_user')
        user_folder_path = os.path.join(MAYER_FOLDER_PATH, user_name)
        
        if not os.path.exists(user_folder_path):
            os.makedirs(user_folder_path, exist_ok=True)

        file_id = update.message.photo[-1].file_id if update.message.photo else update.message.document.file_id
        file = await context.bot.get_file(file_id)
        file_path = os.path.join(user_folder_path, file.file_path.split('/')[-1])
        
        await file.download_to_drive(file_path)
        
        await update.message.reply_text(
            'Благодарю! Прошу Вас написать прописку в Украине (почтовый индекс обязателен). \nПример: Ukrajina, Kijev, úľ. Krajská 211/21, 36023'
        )
        return WAITING_FOR_SLOVAKIA_REGISTRATION
    else:
        await update.message.reply_text(
            'Пожалуйста, отправьте скан Беженства/ВНЖ в формате фотографии или документа.'
        )
        return WAITING_FOR_REFUGEE

# Обработка прописки в Словакии
async def handle_slovakia_registration(update: Update, context: CallbackContext) -> int:
    slovakia_registration = update.message.text.strip()
    user_name = context.user_data.get('full_name', 'unknown_user')

    # Обновление данных в Excel
    df = pd.read_excel(EXCEL_FILE_PATH)
    df.loc[df['Имя Работника'] == user_name, 'Прописка в Словакии'] = slovakia_registration
    df.to_excel(EXCEL_FILE_PATH, index=False)
    adjust_column_widths(EXCEL_FILE_PATH)
    adjust_row_heights_and_alignments(EXCEL_FILE_PATH)

    await update.message.reply_text(
        'Спасибо! Прошу Вас написать прописку в Словакии (почтовый индекс обязателен)\nПример: Slovensko, Bratislava, úľ. Krajská  211/21, 36023'
    )
    return WAITING_FOR_UKRAINE_REGISTRATION

# Обработка прописки в Украине
async def handle_ukraine_registration(update: Update, context: CallbackContext) -> int:
    ukraine_registration = update.message.text.strip()
    user_name = context.user_data.get('full_name', 'unknown_user')

    # Обновление данных в Excel
    df = pd.read_excel(EXCEL_FILE_PATH)
    df.loc[df['Имя Работника'] == user_name, 'Прописка в Украине'] = ukraine_registration
    df.to_excel(EXCEL_FILE_PATH, index=False)
    adjust_column_widths(EXCEL_FILE_PATH)
    adjust_row_heights_and_alignments(EXCEL_FILE_PATH)

    # Настройка ширины столбцов с помощью openpyxl
    from openpyxl import load_workbook
    workbook = load_workbook(EXCEL_FILE_PATH)
    worksheet = workbook.active
    column_widths = {
        'A': 30,  # Имя Работника
        'B': 40,  # Прописка в Украине
        'C': 40,  # Прописка в Словакии
        'D': 30,  # IBAN
        'E': 20,  # Номер телефона
        'F': 15,  # Страховка
        'G': 50,  # Информация о страховке
        'H': 30,  # Вторая работа
        'I': 50,  # Информация о второй работе
        'J': 20,  # Образование
    }
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()
    
    await update.message.reply_text(
        'Спасибо! Попрошу вас отправить свой IBAN\nПример: SK00 0000 0000 0000 0000 0000'
    )
    return WAITING_FOR_BANK_DETAILS

async def handle_bank_details(update: Update, context: CallbackContext) -> int:
    iban = update.message.text.strip()
    user_name = context.user_data.get('full_name', 'unknown_user')

    # Обновление данных в Excel
    df = pd.read_excel(EXCEL_FILE_PATH)
    df.loc[df['Имя Работника'] == user_name, 'IBAN'] = iban
    df.to_excel(EXCEL_FILE_PATH, index=False)
    adjust_column_widths(EXCEL_FILE_PATH)
    adjust_row_heights_and_alignments(EXCEL_FILE_PATH)
    
    await update.message.reply_text(
        'Благодарю, попрошу вас отправить свой словацкий/украинский актуальный номер телефона.\nПример: +421951355537 или +380939843417'
    )
    return WAITING_FOR_SLOVAK_PHONE

# Обработка номера телефона
async def handle_slovak_phone(update: Update, context: CallbackContext) -> int:
    phone_number = update.message.text.strip()
    user_name = context.user_data.get('full_name', 'unknown_user')

    # Обновление данных в Excel
    df = pd.read_excel(EXCEL_FILE_PATH)
    df.loc[df['Имя Работника'] == user_name, 'Номер телефона'] = phone_number
    df.to_excel(EXCEL_FILE_PATH, index=False)
    adjust_column_widths(EXCEL_FILE_PATH)
    adjust_row_heights_and_alignments(EXCEL_FILE_PATH)

    # Создаем клавиатуру с вариантами "да" и "нет"
    reply_keyboard = [['да', 'нет']]
    await update.message.reply_text(
        'Спасибо! У вас имеется страховка?',
        reply_markup=ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=True)
    )
    return WAITING_FOR_INSURANCE


# Обработка информации о страховке
async def handle_insurance(update: Update, context: CallbackContext) -> int:
    user_input = update.message.text.strip().lower()
    user_name = context.user_data.get('full_name', 'unknown_user')

    # Обновление данных в Excel
    df = pd.read_excel(EXCEL_FILE_PATH)
    df.loc[df['Имя Работника'] == user_name, 'Страховка'] = user_input.capitalize()
    df.to_excel(EXCEL_FILE_PATH, index=False)
    adjust_column_widths(EXCEL_FILE_PATH)
    adjust_row_heights_and_alignments(EXCEL_FILE_PATH)
    
    if user_input == 'да':
        # Переходим к запросу информации о страховке
        await update.message.reply_text(
            'Пожалуйста, укажите данные о вашей страховке.'
        )
        return WAITING_FOR_INSURANCE_INFO
    elif user_input == 'нет':
        # Переходим к запросу информации о втором месте работы
        await update.message.reply_text(
            'Есть ли у вас другая работа?'
        )
        return WAITING_FOR_SECOND_JOB_CONFIRMATION
    else:
        await update.message.reply_text(
            'Пожалуйста, выберите "да" или "нет".'
        )
        return WAITING_FOR_INSURANCE


# Обработка информации о страховке (детали)
async def handle_insurance_info(update: Update, context: CallbackContext) -> int:
    user_input = update.message.text.strip()
    user_name = context.user_data.get('full_name', 'unknown_user')

    # Обновление данных в Excel
    df = pd.read_excel(EXCEL_FILE_PATH)
    df.loc[df['Имя Работника'] == user_name, 'Информация о страховке'] = user_input
    df.to_excel(EXCEL_FILE_PATH, index=False)
    adjust_column_widths(EXCEL_FILE_PATH)
    adjust_row_heights_and_alignments(EXCEL_FILE_PATH)
    
    # Переход к следующему шагу (например, второй работе)
    await update.message.reply_text(
        'Спасибо! Есть ли у вас другая работа?'
    )
    return WAITING_FOR_SECOND_JOB_CONFIRMATION


# Обработка подтверждения второго места работы
async def handle_second_job_confirmation(update: Update, context: CallbackContext) -> int:
    user_input = update.message.text.strip().lower()
    if user_input == 'да':
        # Переходим к запросу второго места работы
        await update.message.reply_text(
            'Пожалуйста, укажите второе место работы.'
        )
        return WAITING_FOR_SECOND_JOB
    elif user_input == 'нет':
        # Переходим к запросу информации об образовании
        await update.message.reply_text(
            'Вы еще учитесь?',
            reply_markup=ReplyKeyboardMarkup([['да', 'нет']], one_time_keyboard=True)
        )
        return WAITING_FOR_EDUCATION_CONFIRMATION
    else:
        await update.message.reply_text(
            'Пожалуйста, выберите "да" или "нет".'
        )
        return WAITING_FOR_SECOND_JOB_CONFIRMATION

# Обработка информации о втором месте работы
async def handle_second_job(update: Update, context: CallbackContext) -> int:
    second_job = update.message.text.strip()
    user_name = context.user_data.get('full_name', 'unknown_user')
    
    # Обновление данных в Excel
    df = pd.read_excel(EXCEL_FILE_PATH)
    df.loc[df['Имя Работника'] == user_name, 'Вторая работа'] = second_job  # Записываем ответ в столбец "Вторая работа"
    df.loc[df['Имя Работника'] == user_name, 'Информация о второй работе'] = second_job  # Также записываем ответ в столбец "Информация о второй работе"
    df.to_excel(EXCEL_FILE_PATH, index=False)
    adjust_column_widths(EXCEL_FILE_PATH)
    adjust_row_heights_and_alignments(EXCEL_FILE_PATH)
    
    await update.message.reply_text(
        'Спасибо! Вы еще учитесь?',
        reply_markup=ReplyKeyboardMarkup([['да', 'нет']], one_time_keyboard=True)
    )
    return WAITING_FOR_EDUCATION_CONFIRMATION




# Обработка подтверждения обучения
async def handle_education_confirmation(update: Update, context: CallbackContext) -> int:
    user_input = update.message.text.strip().lower()
    user_name = context.user_data.get('full_name', 'unknown_user')
    
    # Обновление данных в Excel
    df = pd.read_excel(EXCEL_FILE_PATH)
    df.loc[df['Имя Работника'] == user_name, 'Образование'] = 'Учится' if user_input == 'да' else 'Не учится'
    df.to_excel(EXCEL_FILE_PATH, index=False)
    adjust_column_widths(EXCEL_FILE_PATH)
    adjust_row_heights_and_alignments(EXCEL_FILE_PATH)
    
    if user_input == 'да':
        # Переходим к запросу фото студенческого билета
        await update.message.reply_text(
            'Пожалуйста, отправьте фото студенческого билета.'
        )
        return WAITING_FOR_STUDENT_CARD
    elif user_input == 'нет':
        # Переходим к запросу информации об образовании
        await update.message.reply_text(
            'Укажите ваше образование.'
        )
        return WAITING_FOR_EDUCATION
    else:
        await update.message.reply_text(
            'Пожалуйста, выберите "да" или "нет".'
        )
        return WAITING_FOR_EDUCATION_CONFIRMATION

# Обработка информации об образовании
async def handle_education(update: Update, context: CallbackContext) -> int:
    user_input = update.message.text.strip()
    user_name = context.user_data.get('full_name', 'unknown_user')
    
    # Обновление данных в Excel
    df = pd.read_excel(EXCEL_FILE_PATH)
    df.loc[df['Имя Работника'] == user_name, 'Образование'] = user_input
    df.to_excel(EXCEL_FILE_PATH, index=False)
    adjust_column_widths(EXCEL_FILE_PATH)
    adjust_row_heights_and_alignments(EXCEL_FILE_PATH)
    
    
    await update.message.reply_text(
        'Благодарим за предоставленную информацию, с вами свяжется наш менеджер Арсений🧑‍💼\n+421951831134 ☎️'
    )
    return ConversationHandler.END

# Обработка фото студенческого билета
async def handle_student_card(update: Update, context: CallbackContext) -> int:
    if update.message.photo or update.message.document:
        user_name = context.user_data.get('full_name', 'unknown_user')
        user_folder_path = os.path.join(MAYER_FOLDER_PATH, user_name)
        
        # Проверка, что папка существует
        if not os.path.exists(user_folder_path):
            os.makedirs(user_folder_path, exist_ok=True)

        # Получаем file_id
        file_id = update.message.photo[-1].file_id if update.message.photo else update.message.document.file_id

        file = await context.bot.get_file(file_id)
        file_path = os.path.join(user_folder_path, file.file_path.split('/')[-1])
        
        await file.download_to_drive(file_path)
        
        await update.message.reply_text(
            'Спасибо! Теперь отправьте подтверждение из вуза.'
        )
        return WAITING_FOR_UNIVERSITY_CONFIRMATION
    else:
        await update.message.reply_text(
            'Пожалуйста, отправьте фото студенческого билета в формате фотографии или документа.'
        )
        return WAITING_FOR_STUDENT_CARD

# Обработка подтверждения из вуза
async def handle_university_confirmation(update: Update, context: CallbackContext) -> int:
    if update.message.photo or update.message.document:
        user_name = context.user_data.get('full_name', 'unknown_user')
        user_folder_path = os.path.join(MAYER_FOLDER_PATH, user_name)
        
        # Проверка, что папка существует
        if not os.path.exists(user_folder_path):
            os.makedirs(user_folder_path, exist_ok=True)

        # Получаем file_id
        file_id = update.message.photo[-1].file_id if update.message.photo else update.message.document.file_id

        file = await context.bot.get_file(file_id)
        file_path = os.path.join(user_folder_path, file.file_path.split('/')[-1])
        
        await file.download_to_drive(file_path)
        
        await update.message.reply_text(
            'Благодарим за предоставленную информацию, с вами свяжется наш менеджер Арсений🧑‍💼 \n+421951831134 ☎️'
        )
        return ConversationHandler.END
    else:
        await update.message.reply_text(
            'Пожалуйста, отправьте подтверждение из вуза в формате фотографии или документа.'
        )
        return WAITING_FOR_UNIVERSITY_CONFIRMATION

# Обработка неожиданных сообщений
async def handle_unknown(update: Update, context: CallbackContext) -> None:
    await update.message.reply_text(
        'Пожалуйста, следуйте инструкциям и предоставьте запрашиваемую информацию.'
    )

# Основная функция, которая запускает бота
def main() -> None:
    application = Application.builder().token(TOKEN).build()

    # Настройка ConversationHandler
    conversation_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            WAITING_FOR_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_name)],
            WAITING_FOR_PASSPORT: [MessageHandler(filters.PHOTO | filters.Document.ALL, handle_passport)],
            WAITING_FOR_REFUGEE: [MessageHandler(filters.PHOTO | filters.Document.ALL, handle_refugee)],
            WAITING_FOR_SLOVAKIA_REGISTRATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_slovakia_registration)],
            WAITING_FOR_UKRAINE_REGISTRATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_ukraine_registration)],
            WAITING_FOR_BANK_DETAILS: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_bank_details)],
            WAITING_FOR_SLOVAK_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_slovak_phone)],
            WAITING_FOR_INSURANCE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_insurance)],
            WAITING_FOR_INSURANCE_INFO: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_insurance_info)],
            WAITING_FOR_SECOND_JOB_CONFIRMATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_second_job_confirmation)],
            WAITING_FOR_SECOND_JOB: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_second_job)],
            WAITING_FOR_EDUCATION_CONFIRMATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_education_confirmation)],
            WAITING_FOR_EDUCATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_education)],
            WAITING_FOR_STUDENT_CARD: [MessageHandler(filters.PHOTO | filters.Document.ALL, handle_student_card)],
            WAITING_FOR_UNIVERSITY_CONFIRMATION: [MessageHandler(filters.PHOTO | filters.Document.ALL, handle_university_confirmation)],
        },
        fallbacks=[MessageHandler(filters.ALL & ~filters.COMMAND, handle_unknown)]
    )

    application.add_handler(conversation_handler)

    # Запускаем бота
    application.run_polling()

if __name__ == '__main__':
    main()
